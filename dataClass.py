from pathlib import Path
from scraping import scrap , Site
from constants import RESULTS
from import_logging import get_logger
import pandas 
import shutil
import os
from  stat import S_IWRITE
from openpyxl import load_workbook
from time import sleep
import random  
import UI
import sqlite3
import json
from sortedcontainers import SortedList
DB_PATH = "data.db"
ASAN_FOLDER = Path("input asan7")
ASAN_CODES_DICT = Path("Asan Codes Dictionary") # for new products must update: without bug and duplicates
INPUT_FOLDER = Path("input xlsx")
OUTPUT_FOLDER = Path("output xlsx")

logger = get_logger(__name__)

class Data :
    def __init__(self, id, name, price , active, quantity, asan=None):
        self.id = int(id)
        self.name = str(name)
        self.price = int(price)
        self.active = bool(active)
        self.quantity = int(quantity)
        self.asa = asan
        self.sites = None
        self.chosen_site = None
        self.torob_url = None
        self.new_quantity = self.quantity

    def __repr__(self):
        return (f"id='{self.id}, name={self.name!r}, price={self.price}, quantity={self.quantity}, new_quantity={self.new_quantity}, asa={self.asa}, chosen_site={self.chosen_site!r}, scraped?={bool(self.torob_url)}, active={self.active})")

    def update(self):
        "update the product best sites price from trob and return (ready to use in ui )queue of it "
        logger.info(f"Scraping for product (id='{self.id}')...")
        sites , self.torob_url = scrap(self.name , self.torob_url) 
        while len(sites) < RESULTS:
            sites.add(Site(shop_name=None , price=0, badged=False))
        
        for site in sites[:RESULTS]:
            if "اسپارک دیجی" == site.name:
                self.sites = sites[:RESULTS]
                break
        else:
            self.sites= SortedList(sites[:RESULTS-1])
            self.sites.add(Site("اسپارک دیجی", self.price))
        logger.debug(f"list of boxes (updated) to show in ui (len:{len(self.sites)}): {self.sites}")


class Asan:
    def __init__(self ,quantity ,fee ,last_buyed):
        self.quantity = quantity
        self.fee = fee
        self.last_buyed = last_buyed

    def __repr__(self):
        return (f"<quantity={self.quantity}, fee={self.fee}, last_buyed={self.last_buyed}>")


class DataDB:
    def __init__(self, db_path=DB_PATH):
        self.conn = sqlite3.connect(str(db_path))
        self.cursor = self.conn.cursor()
        self._create_table()
        self._create_translation_table()
    
    def _create_table(self):
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS data (
                id           INTEGER PRIMARY KEY,
                name         TEXT    NOT NULL,
                price        INTEGER NOT NULL,
                active       BOOLEAN NOT NULL,
                quantity     INTEGER NOT NULL,
                new_quantity INTEGER,
                sites        TEXT,
                chosen_site  TEXT,
                torob_url    TEXT,
                asa          TEXT
            )
        """)
        self.conn.commit()

    def load_all(self):
        self.cursor.execute("""
            SELECT id, name, price, active, quantity, new_quantity, sites, chosen_site, torob_url, asa
            FROM data
        """)
        rows = self.cursor.fetchall()
        result = []
        for id_, name, price, active, quantity, new_quantity, sites_json, chosen_site, torob_url, asa_json in rows:
            d = Data(id_, name, price, active, quantity)
            d.new_quantity = new_quantity
            if asa_json:
                asa_dict = json.loads(asa_json)
                d.asa = Asan(asa_dict.get("quantity"),asa_dict.get("fee"),asa_dict.get("last_buyed"))
            if sites_json:
                raw_sites = json.loads(sites_json)
                d.sites = [Site.from_dict(s) for s in raw_sites]
            else:
                d.sites = None
            if isinstance(chosen_site, str) and len(chosen_site)!=1 :
                d.chosen_site = int(chosen_site)
            else :
                d.chosen_site = chosen_site
            d.torob_url   = torob_url
            result.append(d)
        return result

    def save_all(self, data_list):
        self.cursor.execute("DELETE FROM data")
        payload = []
        for d in data_list:
            sites_json = None if d.sites is None else json.dumps([s.to_dict() for s in d.sites])
            asa_json = None if d.asa is None else json.dumps({"quantity": d.asa.quantity,"fee": d.asa.fee,"last_buyed": d.asa.last_buyed})
            payload.append((d.id, d.name, d.price, d.active, d.quantity, d.new_quantity, sites_json, d.chosen_site, d.torob_url, asa_json))
        self.cursor.executemany("""
            INSERT OR REPLACE INTO data
            (id, name, price, active, quantity, new_quantity, sites, chosen_site, torob_url, asa)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, payload)
        self.conn.commit()

    def update_chosen(self, d):
        self.cursor.execute("""
            UPDATE data
            SET chosen_site = ?
            WHERE id = ?
        """, (d.chosen_site, d.id))
        self.conn.commit()

    def update_quantity(self, d):
        self.cursor.execute("""
            UPDATE data
            SET new_quantity = ?
            WHERE id = ?
        """, (d.new_quantity, d.id))
        self.conn.commit()

    def update(self, d):
        sites_json = json.dumps([s.to_dict() for s in (d.sites or [])])
        self.cursor.execute("""
            UPDATE data
            SET sites = ?,
                torob_url = ?
            WHERE id = ?
        """, (sites_json, d.torob_url, d.id))
        self.conn.commit()

    def _create_translation_table(self):
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS translations (
                code INTEGER PRIMARY KEY,
                id INTEGER NOT NULL
            )
        """)
        self.conn.commit()

    def load_translation_data(self):
        self.cursor.execute("""
            SELECT code, id
            FROM translations
        """)
        rows = self.cursor.fetchall()
        return {code: id for code, id in rows}

    def save_translation_data(self, data):

        self.cursor.execute("DELETE FROM translations")
        payload = data.items()
        self.cursor.executemany("""
            INSERT INTO translations (code, id)
            VALUES (?, ?)
        """, payload)
        self.conn.commit()

    def close(self):
        self.conn.close()


def asan_file(id_code_map : dict[int, int]):
    """Load Asan data from Excel file and translate it"""
    try:
        asan_path = next(ASAN_FOLDER.glob("*.xlsx"))
        df_asan = pandas.read_excel(asan_path)
    except (StopIteration, FileNotFoundError, OSError) as e:
        logger.error(f"asan_file: XLSX file not found: {e}")
        return {}

    asan_list = {}
    report =[]
    for _, row in df_asan.iterrows():
        code = str(row["کد کالا"])
        id = id_code_map.get(code)
        if id:
            asan_list[id] = Asan(row["مقداراصلی"], row["فی فروش  1"], row["آخرین خرید"])
        report.append(f"{code} : {id}")
    logger.debug("id:code dict= " + " | ".join(report))
    return asan_list


class DataList :
    """
        creat one object of this class for getting a Data list from xlsx and use it\n
        class contains a list[Data]
    """
    __index = -1
    __list_data : list[Data]= []
    def __init__(self , db , re_do) :
        self.__db = db
        try:
            mojodi_asan = asan_file(self.__db.load_translation_data())
            xlsx_file_path = next(INPUT_FOLDER.glob("*.xlsx"))
            # Copy xlsx file to output folder
            if OUTPUT_FOLDER.exists():
                shutil.rmtree(OUTPUT_FOLDER , onerror=lambda func, path, _: (os.chmod(path, S_IWRITE), func(path)))
            OUTPUT_FOLDER.mkdir(exist_ok=True)
            self.output_path = OUTPUT_FOLDER / xlsx_file_path.name
            shutil.copy(xlsx_file_path, self.output_path)
            # reading xlsx file
            df = pandas.read_excel(xlsx_file_path)
            
            active_products = df.groupby("id_product")["active"].apply(lambda a: (a == 1).any())
            active_products = active_products[active_products].index
            df_active = df["id_product"].isin(active_products)

            quantity_default_rows = df[df["default_on"] == 1]
            product_quantitys = quantity_default_rows.set_index("id_product")["quantity"]
            availables_products = product_quantitys[product_quantitys > 0].index         
            df_availables =df["id_product"].isin(availables_products)

            df_mojodi_asan = df["id_product"].isin(mojodi_asan.keys())
            
            # saving self.output_df to "ram" for using in saveData :
            self.output_df = df[(df_active & df_availables) | df_mojodi_asan]
            if re_do and Path(DB_PATH).exists():
                # trying to acces db 
                self.__list_data = self.__db.load_all()
                logger.info(f"list_data exported from db='{DB_PATH}'")
            else :
                # filteering table to Only select "id_product", "name", "price"   from "active" or "mojodi_asan" rows:
                # hint: ~df["active"].isna()  &  df["active"] == 1   are for creating data just one per rows of product (product have some rows)
                filtered = df.loc[((df["active"] == 1) & df_availables) | (df_mojodi_asan & df["active"].notna()) , ["id_product", "name", "price", "active"]]
                self.__list_data = [Data(row.id_product, row.name, row.price , row.active, product_quantitys.get(row.id_product), mojodi_asan.get(row.id_product)) for row in filtered.itertuples()]
                logger.info(f"list_data exported from xlsx'={xlsx_file_path}")
                self.__db.save_all(self.__list_data)
                logger.important("New __list_data:")  
                for d in self.__list_data:
                    logger.important(f"ID='{d.id}'={d.price} '_' {d.quantity} '_' {d.asa}")
                logger.important("----------------------------------------------------------")  
            self.len = len(self.__list_data)
            logger.debug(f"list_data=(len= {self.len}) {self.__list_data}") 
        except StopIteration:
            e = "No xlsx files found in the 'input xlsx' folder"
            logger.error(e+" → sys.exit(1) called")
            raise FileNotFoundError(e)
        except Exception as e :
            logger.exception(f"error during initing DataList: ")
            raise

    def current(self)-> Data :
        return self.__list_data[self.__index]
    
    def updateAll(self , retry_failures):
        logger.info("→→→→→→→→→ 'All prudacts Updating'")
        bar = UI.ProgressDialog()
        if retry_failures:
            for i, d in enumerate(self.__list_data):
                if d.sites is None or d.sites[1].name is None :
                    d.update()
                    self.__db.update(d)
                canceled = bar.progress(i)
                if canceled:
                    logger.info(f"Canceled!")
                    break
        else:
            for i, d in enumerate(self.__list_data):
                if d.sites is None :
                    d.update()
                    self.__db.update(d)
                canceled = bar.progress(i)
                if canceled:
                    logger.info(f"Canceled!")
                    break
            # sleep(random.uniform(2, 4))
        logger.info(f"→→→→→→→→→ 'All prudacts Updating' : procces ended : '{i}' of '{self.len}' prudacts update called")

    def showData(self , is_next : bool) -> Data:
        """This will go forward in the list and return the next Data object."""
        try:
            if is_next :
                self.__index += 1
            else :    
                self.__index -= 1
            if self.__index < -self.len:
                self.__index = self.len -1
            elif self.__index >= self.len:
                self.__index = 0
            curr = self.current()
            logger.info("----------------------------------------------------------")  
            logger.info(f"Current index of list: {self.__index} --refers to id='{curr.id}'")
            if curr.sites is None:
                curr.update()
                self.__db.update(curr)
            return curr , self.__index
        except Exception :
            logger.exception(f"Unexpected error in showData: ")
            return None

    def saveData(self) :
        """Save all of changes maded untill now from memory to xlsx file"""
        try:
            df = self.output_df
            # filter unchangeds & include changes :
            new_rows = []
            data_map = {d.id: d for d in self.__list_data}
            remain = None
            prev_id = None
            for row in df.itertuples(index=False):
                curr_id = row.id_product
                row_dict = row._asdict()
                if prev_id == curr_id : # combination row
                    if remain :
                        if row_dict["default_on"] == 1 and data.new_quantity != data.quantity:
                            row_dict["quantity"] = data.new_quantity
                            logger.info(f"id='{curr_id}' quantity updated to {data.new_quantity}")
                        new_rows.append(row_dict)
                    continue
                prev_id = curr_id
                data = data_map.get(curr_id)
                chosen = data.chosen_site
                if chosen is None: # not reached before save button 
                    logger.debug(f"id='{curr_id}' deleted from xlsx becuase: not reached before save button")
                    remain = False
                    continue  # Remove
                if isinstance(chosen, str):
                    site = data.sites[int(chosen)]
                    if isinstance(site.suggested_price, str): # "dont change price"
                        remain = False
                    else:
                        row_dict["price"] = site.suggested_price
                        remain = True
                elif isinstance(chosen, int): # custom price 
                    row_dict["price"] = chosen 
                    remain = True
                if remain:
                    logger.info(f"id='{curr_id}' price updated to {row_dict["price"]}")
                    new_rows.append(row_dict)
                elif data.new_quantity != data.quantity:
                    remain = True
                    new_rows.append(row_dict)
                else:
                    logger.debug(f"id='{curr_id}' deleted from xlsx becuase: not any change accepted")
            # Save (without extra changings (header style)) :
            wb = load_workbook(self.output_path)
            ws = wb.active
            ws.delete_rows(2, ws.max_row)  # keep header, remove all rows
            for row_idx, row in enumerate(new_rows, start=2):
                for col_idx, col_name in enumerate(df.columns, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))
            wb.save(self.output_path)
            logger.info(f"Successfully new xlsx saved to {self.output_path}")
        except Exception as e:
            logger.exception("Error during saving data to xlsx: ")
            raise e


if __name__ == "__main__":
    print("this file is only class mudule and can not be runed")